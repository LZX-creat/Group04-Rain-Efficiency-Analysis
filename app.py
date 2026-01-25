from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from datetime import timedelta
import csv
import io
import os
from flask_sqlalchemy import SQLAlchemy
import pymysql
import openpyxl  # 在文件顶部添加引用
from flask import send_file

# 必须在初始化前调用
pymysql.install_as_MySQLdb()

app = Flask(__name__)
app.secret_key = 'abc123'
app.config.update(
    SESSION_COOKIE_PATH='/',
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_HTTPONLY=True,
)

# --- MySQL 8 配置 ---
DB_USER = "root"
DB_PASS = "gwm081897"
DB_HOST = "localhost"
DB_NAME = "rain_system"

app.config['SQLALCHEMY_DATABASE_URI'] = f"mysql://{DB_USER}:{DB_PASS}@{DB_HOST}/{DB_NAME}?charset=utf8mb4"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
db = SQLAlchemy(app)


# --- 数据库模型 ---
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default='操作员')


class RainData(db.Model):
    __tablename__ = 'rain_data'
    id = db.Column(db.Integer, primary_key=True)
    year = db.Column(db.Integer, nullable=False)
    total = db.Column(db.Float, nullable=False)
    overflow = db.Column(db.Float, nullable=False)
    car_wash = db.Column(db.Float, default=0.0)
    irrigation = db.Column(db.Float, default=0.0)


# --- 自动建表与初始化 ---
with app.app_context():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', password='admin123', role='管理员')
        db.session.add(admin)
        db.session.commit()


# --- 登录权限验证 ---
@app.before_request
def require_login():
    if request.path.startswith('/static') or request.path == '/favicon.ico':
        return
    white_list = ['/login_page', '/api/login']
    user_logged_in = 'user' in session
    if user_logged_in and request.path == '/login_page':
        return redirect(url_for('index'))
    if not user_logged_in and request.path not in white_list:
        if request.path.startswith('/api/'):
            return jsonify({"status": "error", "message": "请登录"}), 401
        return redirect(url_for('login_page'))


# --- 登录 API ---
@app.route('/api/login', methods=['POST'])
def api_login():
    username = request.form.get('username')
    password = request.form.get('password')
    # 从数据库查找用户
    user = User.query.filter_by(username=username, password=password).first()
    if user:
        session.clear()
        session['user'] = {'username': user.username, 'role': user.role}
        session.permanent = True
        return jsonify({"status": "success", "message": "登录成功"})
    return jsonify({"status": "error", "message": "账号或密码错误"}), 401


# --- 基础页面路由 ---
@app.route('/')
def index(): return render_template('index.html')


@app.route('/login_page')
def login_page(): return render_template('login.html')


@app.route('/logout_page')
def logout_page(): return render_template('logout.html')


@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({"status": "success"})


@app.route('/stats')
def stats(): return render_template('stats.html')


# --- 新增：模板下载接口 ---
@app.route('/api/download_template')
def download_template():
    # 创建内存中的 Excel 文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "雨水数据导入模板"

    # 定义表头
    headers = ["年份", "降雨量(mm)", "溢流量(mm)", "洗车利用量(m³)", "灌溉利用量(m³)"]
    ws.append(headers)

    # 填充一行示例数据（可选）
    ws.append([2024, 1050.5, 210.2, 50.0, 120.0])

    # 保存到内存流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='rain_data_template.xlsx'
    )

@app.route('/settings')
def settings(): return render_template('settings.html')


@app.route('/import_page')
def import_page(): return render_template('import.html')


# --- 账号管理 API ---
@app.route('/api/users')
def get_users():
    users = User.query.all()
    return jsonify([{"id": u.id, "username": u.username, "password": u.password, "role": u.role} for u in users])


@app.route('/api/users/add', methods=['POST'])
def add_user():
    username = request.form['username']
    if User.query.filter_by(username=username).first():
        return jsonify({"status": "error", "message": "用户名已存在"}), 400
    new_user = User(
        username=username,
        password=request.form['password'],
        role=request.form['role']
    )
    db.session.add(new_user)
    db.session.commit()
    return jsonify({"status": "success"})


@app.route('/api/users/edit', methods=['POST'])
def edit_user():
    user = User.query.get(request.form['id'])
    if user:
        user.username = request.form['username']
        user.password = request.form['password']
        user.role = request.form['role']
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404


@app.route('/api/users/delete/<int:id>', methods=['POST'])
def delete_user(id):
    user = User.query.get(id)
    if user:
        db.session.delete(user)
        db.session.commit()
    return jsonify({"status": "success"})


# --- 降雨数据 API ---
@app.route('/api/data')
def get_data():
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 5, type=int)
    search_year = request.args.get('year', '')
    rain_min = request.args.get('min', '')
    rain_max = request.args.get('max', '')

    query = RainData.query
    if search_year:
        query = query.filter(RainData.year == int(search_year))
    if rain_min:
        query = query.filter(RainData.total >= float(rain_min))
    if rain_max:
        query = query.filter(RainData.total <= float(rain_max))

    # 计算分页
    total_records = query.count()
    total_pages = (total_records + page_size - 1) // page_size

    # 应用分页
    pagination = query.order_by(RainData.year.desc()) \
        .offset((page - 1) * page_size) \
        .limit(page_size) \
        .all()

    records = []
    t_total, t_overflow, t_reuse = 0, 0, 0

    for r in pagination:
        rate = ((r.total - r.overflow) / r.total) if r.total > 0 else 0
        reuse = r.car_wash + r.irrigation
        records.append({
            "id": r.id, "year": r.year, "total": r.total, "overflow": r.overflow,
            "car_wash": r.car_wash, "irrigation": r.irrigation,
            "rate_num": round(rate * 100, 2),
            "rate_str": f"{rate * 100:.2f}%"
        })
        t_total += r.total
        t_overflow += r.overflow
        t_reuse += reuse

    count = len(records)
    avg_rate = ((t_total - t_overflow) / t_total * 100) if t_total > 0 else 0

    return jsonify({
        "records": records,
        "total": total_records,  # 总记录数
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "stats": {
            "count": count,
            "avg_rain": round(t_total / count, 2) if count > 0 else 0,
            "avg_rate": round(avg_rate, 2),
            "total_reuse": round(t_reuse, 2)
        }
    })


# --- 新增：专门用于图表的全量数据接口 ---
@app.route('/api/all_stats')
def get_all_stats():
    # 获取所有记录，按年份升序排列（图表通常需要从旧到新）
    all_records = RainData.query.order_by(RainData.year.asc()).all()

    records = []
    t_total, t_overflow, t_reuse = 0, 0, 0

    for r in all_records:
        rate = ((r.total - r.overflow) / r.total) if r.total > 0 else 0
        reuse = r.car_wash + r.irrigation
        records.append({
            "year": r.year,
            "total": r.total,
            "overflow": r.overflow,
            "car_wash": r.car_wash,
            "irrigation": r.irrigation,
            "rate_num": round(rate * 100, 2)
        })
        t_total += r.total
        t_overflow += r.overflow
        t_reuse += reuse

    count = len(records)
    avg_rate = ((t_total - t_overflow) / t_total * 100) if t_total > 0 else 0

    return jsonify({
        "records": records,
        "stats": {
            "avg_rain": round(t_total / count, 2) if count > 0 else 0,
            "avg_rate": round(avg_rate, 2),
            "total_reuse": round(t_reuse, 2)
        }
    })

@app.route('/api/add', methods=['POST'])
def add():
    new_record = RainData(
        year=int(request.form['year']),
        total=float(request.form['total']),
        overflow=float(request.form['overflow']),
        car_wash=float(request.form.get('car_wash') or 0),
        irrigation=float(request.form.get('irrigation') or 0)
    )
    db.session.add(new_record)
    db.session.commit()
    return jsonify({"status": "success"})


@app.route('/api/edit/<int:id>', methods=['POST'])
def edit(id):
    record = RainData.query.get(id)
    if record:
        record.total = float(request.form['total'])
        record.overflow = float(request.form['overflow'])
        record.car_wash = float(request.form['car_wash'])
        record.irrigation = float(request.form['irrigation'])
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404


@app.route('/api/delete/<int:id>', methods=['POST'])
def delete(id):
    record = RainData.query.get(id)
    if record:
        db.session.delete(record)
        db.session.commit()
    return jsonify({"status": "success"})


import openpyxl  # 在文件顶部添加引用


@app.route('/api/upload', methods=['POST'])
def upload_file():
    file = request.files.get('file')
    if not file:
        return jsonify({"status": "error", "message": "无文件"}), 400

    try:
        # 使用 openpyxl 读取内存中的文件流
        wb = openpyxl.load_workbook(file)
        sheet = wb.active  # 获取第一个工作表

        count = 0
        # 假设第一行是表头，从第二行开始读取
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # row 结构: (年份, 总降雨, 溢流, 洗车, 灌溉)
            if not row or row[0] is None:
                continue

            record = RainData(
                year=int(row[0]),
                total=float(row[1] or 0),
                overflow=float(row[2] or 0),
                car_wash=float(row[3] or 0) if len(row) > 3 else 0,
                irrigation=float(row[4] or 0) if len(row) > 4 else 0
            )
            db.session.add(record)
            count += 1

        db.session.commit()
        return jsonify({"status": "success", "message": f"成功从 Excel 导入 {count} 条数据"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"解析失败: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)